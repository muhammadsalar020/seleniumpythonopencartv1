import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    workbook.close()
    return rows


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    cols = sheet.max_column
    workbook.close()
    return cols


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rownum, column=columnno).value
    workbook.close()
    return data


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
    workbook.close()


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')

    sheet.cell(row=rownum, column=columnno).fill = greenFill

    workbook.save(file)
    workbook.close()


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')

    sheet.cell(row=rownum, column=columnno).fill = redFill

    workbook.save(file)
    workbook.close()